#参考Site
#https://qiita.com/sky_jokerxx/items/dc9d8827d946b467ba4b


#ライブラリのimport
import xlrd
import openpyxl
import pandas
import pathlib
import pprint

#対象フォルダの指定
path = pathlib.Path('C:/Users\山本厚士/Dropbox/営業技術課/点検時データ・一時保管場所')
#保存ファイルの指定
saveFile = openpyxl.load_workbook('C:/Users/山本厚士/SkyDrive/DNS/Python/PVT統計.xlsx')

#対象ファイルリスト取得(*.xlsx)
filelist = list(path.glob("**/*.xlsx"))
dataframeList = []

for fileDir in filelist:
    #ファイル名
    fileName = fileDir.name
    #RT-3点検ファイルの検査結果を抽出
    if "RT-3点検" in fileName:
        #Excel FileOpen
        print('ファイル名:%s' % fileName)
        workbook = xlrd.open_workbook(fileDir)
        #workBook指定
        for loadSheet in workbook.sheets():
            #シートの展開
            loadSheetName = loadSheet.name
            #サンプリング液量検査結果
            if loadSheetName == "サンプリング":
                saveSheet = saveFile[loadSheetName]
#                dataframeList.append(saveFile.parse(sheetName))
#                print(saveSheet)
                saveSheet["B1"] = "ABC"
                saveSheet.cell(row=2, column=2, value="DEF")
                saveFile.save('C:/Users/山本厚士/SkyDrive/DNS/Python/PVT統計.xlsx')
                #セル情報の展開
 #               dataframeList.insert(0,'A')
  #              dataframeList.insert(1,'B')
#                print(dataframeList)

                for rowNum in range(loadSheet.nrows):
                    a = 1
#                    print(loadSheet)
#                   print(rowNum ,'/' , sheet.nrows , sheet.row_values(rowNum))



'''
    sheet = wb.sheet_names()
    #セル情報
    for sheetName in sheet:
        print (sheetName)
        for rowNum in sheet.nrows:
            print (sheetName.row_values(rowNum))
#シートclass
sheets = wb.sheets()
print(type(sheets))

#シート指定
sheet = wb.sheet_by_name('Sheet1')
print(type(sheet))

#セル指定
cell = sheet.cell(1,2)
print(cell)
print(type(cell))

#セル値取得
print(cell.value)
print(sheet.cell_value(1,2))

#列データ取得
col = sheet.col(1)
colVol = sheet.col

#行データ取得
row = sheet.row(1)

#列数
print(sheet.ncols)
#行数
print(sheet.nrows)
'''

'''
for rowNum in range(sheet.nrows):
    print (sheet.row_values(rowNum))
'''

#print([sheet.row_values(x) for x in range(5)])
#pprint.pprint([sheet.row_values(x) for x in range(5)])